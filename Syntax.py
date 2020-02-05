# Python is case sensitive.All variable names should be lower cases.
# True or False start with higher cases.

Name = input('What is your name? ')
# This input function will print this message on the terminal, and then it will wait for the user to enter a value. Whatever the user enters this input function will return.

#======================
#Type conversion
int()
float()
bool()

#======================
# String related
course = 'python for "beginners" '
course = "python's for beginnners"
# Multiple lines of string
course = ''' 
Hi
This is multiple lines of string
'''

course = "Python for beginners"
print(course[0])  # Output: P
print(course[-1]) # the first character from the end Output: s
print(course[0:3]) # return character 0, 1, 2 Output: Pyt
print(course[1:])  # Output: ython for begginners

# Formatted string
first = 'John'
last = 'Smith'
# John[Smith] is a coder
Msg = f'{first}  [{last}] is a coder'
# define formatted string: prefix your string with an f and then use curly braces to dynamically insert values into your string

course = 'hellow world'
len()      # general purpose function built in python
course.upper() # method
course.lower()
course.title() #convert the first character in each word to Uppercase and remaining characters to Lowercase in string and returns new string.
course.find()
course.replace()
course.split(' ') #return a list
'…' in course #return boolean

#===========================
#arthemetic
2 ** 3 #Exponentiation operator

a = 2
b = 1
c = a / b  #we get float type
c = a // b #we get integer

# If
is_hot = True
is_cold = False
if is_hot:
    print(1)
elif is_cold:
    print(2)
else:
    print(3)
    print(4)

# while loop
while condition:
    …
else: …  #if this while loop completes successfully without an immediate

# for loop
for item in 'Python':
    …     # iterate each character in string

for item in range(5, 10, 2)
    print(item)
# Output:
#5
#7
#9

#==========================
#list
list = [1, 2, 3, 4, 5]
print(list[2:])
#Output:
3
4
5

list.append(6)
list.insert(index, object)
list.remove(targeted_item)
list.clear()
list.pop()
list.index(item)
list.count(item)
list.sort()
list.reverse()
print(50 in list)
Output: False

# tuple : special list which cannot be motified
Number = (1, 2, 3)


# unpack
coordinates = (1, 2, 3)
X = coordinates[0]
y = coordinates[1]
z = coordinates[2]
# Shorthand:
X, y, z = coordinates #Same for list

# Dictionaries
customer = {
    "name": "Json",
    "age": 20,
    "is_verified": True
}

print(customer["name"])
print(customer.get("birthday")) # don't report error, return none object
print(customer.get("birthday", "jan 1"))  # add a new key
customer["name"] = "jaak" # update value

#==========================
#function
def greet_user(first_name, last_name):
    print("Hi")
    print("Welcome")


print("start")  #add two lines break to this function
greet_user("Smith", first_name="John") # keyword argument: the order doesn't matter
print("end")
# if you're mixing positional and keyword arguments, you should always use the positional arguments first and then keyword arguments
# if you're dealing with functions that takes multiple numerical values and it's not quite clear what those values represent, use keywords arguments to improve the readability of your code

#===========================
#exceptions
try:
    age = int(input('Age: '))
    income = 100/age
    print(income)
except ZeroDivisionError:
    print("a")
except ValueError:
    print('b')

#============================
#class
class Point:
    def move(self):
        print("move")

    def draw(self):
        print('draw')

point1 = Point()
point1.x = 10   #we can set attributes anywhere in our programs

#constructor
class Point:
    def __init__(self,x,y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print('draw')

point = Point(10,20)

#inheritance
class Mammal:
    def walk(self):
        print("walk")

class Dog(Mammal):
    pass # python doesn't like empty class. We can use pass statement, and that basically means nothing

class Cat(Mammal):
    pass

#=========================
# Modules
#app.py
#converters.py

#in app.py, import the entire module: converters.py   we refer to each file as a module
import converters # without extension .py
print(converters.move())
# import specific functions
from converters import move
print(move()) # without prefixing it with the module name

#==========================
# Packages : directory
    # new package: a file named __init__.py is generated automatically. If we add new directory, we need to add __init__.py manually
# approach 1 : import package.module
    # every time we call the function, we need to prefix it with the package, module name
# approach 2: from package.module import function
#               call function
#             from package import module
#               call module.function

#===========================
# Generate random values
import random  #module
random.random() #random value between 0 and 1
random.randint(a,b) #random int between a and b

members = ["a", 'b', 'c', 'd']
leader = random.choice(members)  #randomly pick a leader from the list

class Dice:
    def roll(self):
        x = random.randint(1, 10)
        y = random.randint(1, 10)
        return x, y  #when you want to return a tuple from a function, you don't have to add parenthesis


d1 = Dice()
print(d1.roll())


#============================
# Files and directories
from pathlib import Path #class
# Absolute path
# Relative path

path = Path('ecommerce')
print(path.exists())
# output: True or False

path = Path('email')
path.mkdir()  # create new one
path.rmdir()  # delete

path = Path()  #current directory

for item in path.glob('*.py'):   #path.glob() return a generated object which you need to iterate  '*' all files
    print(item)

#=====================================
# Excel
import openpyxl as xl
# xl.xxx instead of openpyxl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')   # return a workbook
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)   # They're same
print(cell.value)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Draw a chart
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')   # position

wb.save('transactions2.xlsx')

#==============================
# Machine Learning
# Libraries in python:
    # Numpy: provides a multidimensional array
    # Pandas: a data analysis library which provides a concept called data frame. Data frame is a two dimensional data structure similar to an excel spreadsheet(row and column)
    # MatPlotLib: two dimensional plotting library for creating graphs on plots
    # Scikit-Learn: provides all these common algorithms like decision trees, neural networks and so on.

# Jupyter
# install Anaconda environment
# command 'jupyter notebook' in terminal
# Jupyter shortcuts
    # The activated cell can be either in the edit mode(green) or command mode(blue). Depending on the mode, we have different shortcuts. (Press esc in the edit mode can transit to command mode)
    # Command mode
        # h: the list of all the keyboard shortcuts
        # a: insert a new cell above the current cell
        # b: insert a new cell below the current cell
        # dd: delete the cell
    # Edit mode
        # tab: see all the available methods
        # shift+tab: with the cursor on the name of the method, to see the tool tip that describes what this method does and what parameter it takes
        # command+/: comment
    import pandas as pd
    df = pd.read_csv('vgsales.csv')
    df
    df.shape
    df.describe()
    df.values

# real project
    import pandas as pd
    from sklearn.tree import DecisionTreeClassifier

    music_data = pd.read_csv('music.csv')
    X = music_data.drop(columns=['genre'])  # build a new set which has all the columns of the original except for 'genre' column
    y = music_data['genre']  # column 'genre'

    model = DecisionTreeClassifier()
    model.fit(X, y)   # this method takes 2 data sets - the input set and the output set
    predictions = model.predict([[21, 1], [22, 0]])

# calculating the accuracy
    import pandas as pd
    from sklearn.tree import DecisionTreeClassifier
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import accuracy_score

    music_data = pd.read_csv('music.csv')
    X = music_data.drop(columns=['genre'])
    y = music_data['genre']
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2)  # return a tuple

    model = DecisionTreeClassifier()
    model.fit(X_train, y_train)
    predictions = model.predict(X_test)

    score = accuracy_score(y_test, predictions)
    score  # control + enter : run the current cell without inserting new cell

# Persisting models
    # before: save the model
    import pandas as pd
    from sklearn.tree import DecisionTreeClassifier
    from sklearn.externals import joblib  # has methods for loading and saving models

    music_data = pd.read_csv('music.csv')
    X = music_data.drop(columns=['genre'])
    y = music_data['genre']

    model = DecisionTreeClassifier()
    model.fit(X, y)

    joblib.dump(model, 'music-recommender.joblib')  # the file name of what we want to save model as

    # after: load the model
    import pandas as pd
    from sklearn.tree import DecisionTreeClassifier
    from sklearn.externals import joblib  # has methods for loading and saving models

    model = joblib.load('music-recommender.joblib')  # the file name of what we want to save model as
    predictions = model.predict([[21, 1]])
    predictions

# Visualizing a decision tree
    import pandas as pd
    from sklearn.tree import DecisionTreeClassifier
    from sklearn import tree

    music_data = pd.read_csv('music.csv')
    X = music_data.drop(columns=['genre'])
    y = music_data['genre']

    model = DecisionTreeClassifier()
    model.fit(X, y)

    tree.export_graphviz(model, out_file='music-recommder.dot',
                         feature_names=['age', 'gender'],
                         class_names=sorted(y.unique()),
                         label='all',
                         rounded=True,
                         filled=True)

#======================
#django project : develop a website with python
# pip django==2.1
# create new django project
    #django-admin startproject pyshop .
        #django-admin: django brings a command line utility called django admin. This is a utility or program that we can execute from the command line or terminal
        #startproject: this programs takes many arguments. we use this to start a project named pyshop in the current folder
        #.:means current folder. Without it, this utility will create an extra folder
# python3 manage.py runserver
# python3 manage.py startapp products

# View functions
    # products/views.py
    from django.shortcuts import render
    from django.http import HttpRequest

    def index(request):
        return HttpRequest('Hello World')

# Url mapping - Uniform resource locator (address)
    # under products folder, add new py file named 'urls'
    from django.urls import path
    from . import views

    urlpatterns = [
        path('', views.index)
    ]

    # pyshop/url.py
    from django.contrib import admin
    from django.urls import path, include

    urlpatterns = [
        path('admin/', admin.site.urls),
        path('products/', include('products.urls'))
    ]

# Models: orders, products, customers, reviews
# products/models.py
from django.db import models
class Product(models.Model):
    name = models.CharField(max_length=255)
    price = models.FloatField()
    stock = models.IntegerField()
    image = models.CharField(max_length=2083)  # standard maximum length for url

# Migrations: database
# terminal: python3 manage.py makemigrations
# pyshop package - settings.py - INSTALLED APPS[]
    # goto products/apps.py - class ProductsConfig
    # back to INSTALLED APPS[], add 'products.apps.ProductsConfig' - package.models.class
    # go back to terminal, run the last command - then we can found migrations under productions
    # then run 'python3 manage.py migrate' on terminal  - to add tables in database
    # reload the data file to database

# Admin - /admin
# create user
    # terminal: python3 manage.py createsuperuser
    # username:Jessica pw:1234

# products/admin.py
    from django.contrib import admin
    from .models import Product   # models in current folder. Products is the class added in models.py

    admin.site.register(Product)

    # add a new product on website
        # add image url : go to google, search orange, click Tools, from the usage rights select labeled for reuse

# Tips:
    # modify models and update database

    '''
    1.Delete the sqlite database file (often db.sqlite3) in your django project folder (or wherever you placed it)
    2.Delete everything except __init__.py file from migration folder in all django apps
    3.Make changes in your models (models.py).
    4.Run the command python manage.py makemigrations or python3 manage.py makemigrations
    5.Then run the command python manage.py migrate.
    '''

# Modify the admin display
    # products/admin.py
    from django.contrib import admin
    from .models import Product  # models in current folder. Products is the class added in models.py

    class ProductAdmin(admin.ModelAdmin):  # name convention
        list_display = ('name', 'price', 'stock')   # in corresponding with products/models.py Product class

    admin.site.register(Product, ProductAdmin)  # add new class to the arguments


# Templates
# products/views.py
    from django.shortcuts import render
    from django.http import HttpResponse
    from .models import Product

    def index(request):
        products = Product.objects.all()  # returns all the products we have in the database
        # Product.objects.filter/get/save ()
        return HttpResponse('Hello World')

    def new_product(request):
        return HttpResponse('New Product')

# add a new folder named templates under products folder
# add index.html under templates   tips: type h1, tab
    < h1 > Products < / h1 >
    < ul >
    { %
    for product in products %} < !--template tag in django-->
    < li > {{product.name}}(${{product.price}}) < / li >
    { % endfor %}
    < / ul >
# back to views.py

    from django.shortcuts import render
    from django.http import HttpResponse
    from .models import Product

    def index(request):
        products = Product.objects.all()  # returns all the products we have in the database
        # Product.objects.filter/get/save ()
        return render(request, 'index.html',
                      {'products': products})  # we use this function to render a template

    def new_product(request):
        return HttpResponse('New Product')


# Adding Bootstrap
# google getbootstrap
    # starter template  - copy
# build base.html under templates - paste and modify
    # change <h1> to
    # {% block content %}
    # {% endblock %}
# modify index.html
    { % extends
    'base.html' %}

    { % block
    content %}
    < h1 > Products < / h1 >
    < ul >
    { %
    for product in products %} < !--template tag in django-->
    < li > {{product.name}}(${{product.price}}) < / li >
    { % endfor %}
    < / ul >
    { % endblock %}

# Rendering cards
# index.html
    # instead of using li, we use div
        # type div.row + tag
    # paste 'card' component which was copied from document of bootstrap
    # option+command+L to format the code

    { % extends 'base.html' %}

    { % block content %}
        < h1 > Products < / h1 >
        < div class ="row" >
            { % for product in products %}   <!--template tag in django-->
                < div class ="col" >
                    < div class ="card" style="width: 18rem;" >
                    < img src = "{{ product.image }}" class ="card-img-top" alt="" >
                        < div class ="card-body" >
                            < h5 class ="card-title" > {{product.name}} < / h5 >
                            < p class ="card-text" > ${{product.price}} < / p >
                            < a href = "#" class ="btn btn-primary" > Add to Cart < / a >
                        < / div >
                    < / div >
                < / div >
            { % endfor %}
        < / div >
    { % endblock %}

# Final touches
# add navbar
# copy component and paste to base.html(used for all pages)
# create templates directory under the root folder(the project folder - PyShop), then drag the base.html and drop it to the new folder
# in the main package - pyshop - settings.py
TEMPLATES = [
    {
        'BACKEND': 'django.template.backends.django.DjangoTemplates',
        'DIRS': [
            os.path.join(BASE_DIR, 'templates')  # BASE_DIR is default setting above
        ],
        'APP_DIRS': True,
        'OPTIONS': {
            'context_processors': [
                'django.template.context_processors.debug',
                'django.template.context_processors.request',
                'django.contrib.auth.context_processors.auth',
                'django.contrib.messages.context_processors.messages',
            ],
        },
    },
]

# add padding
# base.html
# type out div.container+ tag
<div class="container">
    {% block content %}
    {% endblock %}
</div>





